import openpyxl
#todo open xls
price_up={'Carlic':3.11,'Celery':1.03,'Lemon':1.44}
wr=openpyxl.load_workbook('produceSales.xlsx')
sheet=wr['Sheet']
#todo find the specal object 
for row in range(2,sheet.max_row):
    productname=sheet.cell(row=row,column=1).value
    if productname in price_up:
        sheet.cell(row=row,column=2).value=price_up[productname]
wr.save('produceSalesbak.xlsx')
#todo update the price of object and save as new one