import openpyxl,sys
from openpyxl.utils import get_column_letter,column_index_from_string
#todo get the par from the order
numb=int(sys.argv[1])
#todo make the title row and column
wb=openpyxl.Workbook()
sheet=wb.active
for i in range(1,numb):
    sheet['A'+str(i+1)]=i
    sheet[get_column_letter(i+1)+'1']=i
    for j in range(1,numb):
        sheet[get_column_letter(i+1)+str(j+1)]=i*j
wb.save('multiplication.xls')

PhoneType,BatterName ,BatterNum  ,LineName ,LineNum