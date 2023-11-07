import openpyxl,os
##0.config database connect
##1.read data from excel
for folderName,subfolders,filenames in os.walk('loadfile'):
    for filename in filenames:
        print('now processing the file '+filename)
       # readpath='loadfile\\'+filename
wb=openpyxl.load_workbook('loadfile\\'+filename)
#print(wb.sheetnames)
for sheetname in wb.sheetnames:
    sheet=wb[sheetname]
    #print(sheet['A1'].value)

##2.config filed mapping and save the mapping to the list
    for i in range(sheet.max_column):

##3.conn database and input the data to the database
    for j in range(sheet.max_column):
        instr=

